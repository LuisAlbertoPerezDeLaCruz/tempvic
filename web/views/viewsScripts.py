# -*- coding: utf-8 -*-
#######################
# viewsScripts.py #
#######################

from ..models import *
from django.shortcuts import render, redirect, get_object_or_404
from web.templatetags.filtrosEspeciales import *
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill
from openpyxl.cell import Cell

# Acciones
NADA = 0
ERROR = 1
CREADO = 2
ACTUALIZADO = 3
REFERENCIADO = 4
ACCION={NADA:'sin accion',ERROR:'error generado',CREADO:'creado',ACTUALIZADO:'actualizado',REFERENCIADO:'referenciado'}

resultados = {'user': NADA, 'phone':NADA, 'userProfile': NADA, 'marca':NADA, 'relacion': NADA, 'saldo': NADA, 'plan':NADA}

def crearUsuariosBatch(request):
    # Offsets
    CORREO = 0
    CLAVE = 0
    PRIMER_NOMBRE = 0
    PRIMER_APELLIDO = 0
    SEGUNDO_NOMBRE = 0
    SEGUNDO_APELLIDO = 0
    TELEFONO = 0
    ALIAS = 0
    MARCA = 0
    CIUDAD = 0
    MUNICIPIO = 0
    PAIS = 0
    CALLE_AVENIDA = 0
    EDIF_CASA = 0
    URBANIZACION = 0
    DISCIPLINA1=0
    DISCIPLINA2=0
    DISCIPLINA3=0
    CREDITOS=0

    faltaInfo = False

    workbook = load_workbook(filename='../../sources/dataZonaFuncional.xlsx')
    grabaEncabezados(workbook)

    for row_cells in workbook['atletas'].iter_rows():
        faltaInfo = False
        resultados['user']=NADA
        resultados['phone']=NADA
        resultados['userProfile']=NADA
        resultados['marca']=NADA
        resultados['relacion']=NADA
        resultados['saldo']= NADA
        for cell in row_cells:
            if cell.row == 1 and cell.value is not None:
                if cell.value.strip() == u'Primer Nombre':
                    PRIMER_NOMBRE = cell.col_idx - 1
                elif cell.value.strip() == u'Primer Apellido':
                    PRIMER_APELLIDO = cell.col_idx - 1
                if cell.value.strip() == u'Segundo Nombre':
                    SEGUNDO_NOMBRE = cell.col_idx - 1
                elif cell.value.strip() == u'Segundo Apellido':
                    SEGUNDO_APELLIDO = cell.col_idx - 1
                elif cell.value.strip() == u'Email':
                    CORREO = cell.col_idx - 1
                elif cell.value.strip() == u'Contraseña':
                    CLAVE = cell.col_idx - 1
                elif cell.value.strip() == u'Alias':
                    ALIAS = cell.col_idx - 1
                elif cell.value.strip() == u'Teléfono':
                    TELEFONO = cell.col_idx - 1
                elif cell.value.strip() == u'Relación con Marca':
                    MARCA = cell.col_idx - 1
                elif cell.value.strip() == u'Ciudad':
                    CIUDAD = cell.col_idx - 1
                elif cell.value.strip() == u'Municipio':
                    MUNICIPIO = cell.col_idx - 1
                elif cell.value.strip() == u'País':
                    PAIS = cell.col_idx - 1
                elif cell.value.strip() == u'Calle o Av.':
                    CALLE_AVENIDA = cell.col_idx - 1
                elif cell.value.strip() == u'Edif. / Casa':
                    EDIF_CASA = cell.col_idx - 1
                elif cell.value.strip() == u'Urbanización':
                    URBANIZACION = cell.col_idx - 1
                elif cell.value.strip() == u'Primera Disciplina':
                    DISCIPLINA1 = cell.col_idx - 1
                elif cell.value.strip() == u'Segunda Disciplina':
                    DISCIPLINA2 = cell.col_idx - 1
                elif cell.value.strip() == u'Tercera Disciplina':
                    DISCIPLINA3 = cell.col_idx - 1
                elif cell.value.strip() == u'Creditos Disponibles':
                    CREDITOS = cell.col_idx - 1
            else:
                if cell.row > 1:
                    correo = trim(row_cells[CORREO].value).lower()
                    clave = trim(row_cells[CLAVE].value)
                    nombre = trim(row_cells[PRIMER_NOMBRE].value)
                    apellido = trim(row_cells[PRIMER_APELLIDO].value)
                    segundoNombre = trim(row_cells[SEGUNDO_NOMBRE].value)
                    segundoApellido = trim(row_cells[SEGUNDO_APELLIDO].value)
                    telefono = soloNumeros(trim(row_cells[TELEFONO].value))
                    alias = trim(row_cells[ALIAS].value)
                    marca = trim(row_cells[MARCA].value)
                    calle = trim(row_cells[CALLE_AVENIDA].value)
                    edificio_casa = trim(row_cells[EDIF_CASA].value)
                    urbanizacion = trim(row_cells[URBANIZACION].value)
                    ciudad = trim(row_cells[CIUDAD].value)
                    municipio = trim(row_cells[MUNICIPIO].value)
                    pais = trim(row_cells[PAIS].value)
                    disciplina1 = trim(row_cells[DISCIPLINA1].value)
                    disciplina2 = trim(row_cells[DISCIPLINA2].value)
                    disciplina3 = trim(row_cells[DISCIPLINA3].value)
                    creditos=int(str(row_cells[CREDITOS].value))
                    if alias == None or alias == '':
                        alias = nombre[0:1].lower() + apellido.lower()
                    if clave == None or clave == '':
                        clave = nombre.lower() + '123'
                    if (correo == '' or clave == '' or nombre == '' or apellido == '' or telefono == '' or alias == '' or marca == ''):
                        faltaInfo = True
                    if (ciudad == '' or municipio == '' or pais == ''):
                        faltaInfo = True
        if not faltaInfo and cell.row>1:
            crearUsuario(correo, clave, nombre, apellido, segundoNombre, segundoApellido,
                         telefono, alias, marca,
                         calle, edificio_casa, urbanizacion, ciudad, municipio, pais,
                         disciplina1,disciplina2,disciplina3,creditos)
            grabaResultados(correo,workbook,resultados)
    grabaFooter(workbook)
    return redirect("/")


def crearUsuario(correo, clave, nombre, apellido, segundoNombre,
                 segundoApellido, telefono, alias, marca,
                 calle, edificio_casa, urbanizacion, ciudad, municipio, pais,
                 disciplina1, disciplina2, disciplina3,creditos):

    try:
        usuarioExiste = User.objects.filter(username=correo)
        if usuarioExiste:
            usuario = User.objects.get(username=correo)
            usuario.email = correo
            usuario.first_name = nombre
            usuario.last_name = apellido
            usuario.set_password(clave)
            usuario.is_active=True
            usuario.save()
            resultados['user']=ACTUALIZADO

        else:
            usuario = User.objects.create(username=correo, password=clave, first_name=nombre, last_name=apellido,
                                          email=correo,
                                          is_active=True)
            usuario.set_password(clave)
            usuario.save()
            resultados['user']=CREADO
    except:
        resultados['user'] = ERROR
        return


    try:
        if UserProfile.objects.filter(u_user=usuario).exists() == False:
            perfil = UserProfile.objects.create(
                u_user=usuario,
                u_alias=alias,
                u_secondname=segundoNombre,
                u_secondlastname=segundoApellido,
                u_telefono=telefono,
                u_calle=calle,
                u_edificioCasa=edificio_casa,
                u_urbanizacion=urbanizacion,
                u_direccion="Edf. " + edificio_casa + ", Cll/Av." + calle + " - Urb." + urbanizacion,
                u_ciudad = Ciudad.objects.get(c_nombre=ciudad),
                u_municipio = Zona.objects.get(z_municipio=municipio),
                u_displinafav1=Disciplina.objects.get(d_nombre=disciplina1),
                u_displinafav2=Disciplina.objects.get(d_nombre=disciplina2),
                u_displinafav3=None,
                u_pais = Pais.objects.get(p_nombre=pais)
            )
            if disciplina3 != '':
                perfil.u_displinafav2 = Disciplina.objects.get(d_nombre=disciplina3)
                perfil.save()
            resultados['userProfile'] = CREADO
        else:
            perfil = UserProfile.objects.get(u_user=usuario)
            u_user = usuario,
            u_alias = alias,
            u_secondname = segundoNombre,
            u_secondlastname = segundoApellido,
            u_telefono = tel,
            u_calle = calle,
            u_edificioCasa = edificio_casa,
            u_urbanizacion = urbanizacion,
            u_direccion = "Edf. " + edificio_casa + ", Cll/Av." + calle + " - Urb." + urbanizacion,
            u_ciudad = Ciudad.objects.get(c_nombre=ciudad),
            u_municipio = Zona.objects.get(z_municipio=municipio),
            u_displinafav1 = Disciplina.objects.get(d_nombre=disciplina1),
            u_displinafav2 = Disciplina.objects.get(d_nombre=disciplina2),
            u_pais = Pais.objects.get(p_nombre=pais)
            if disciplina3 != '':
                perfil.u_displinafav2 = Disciplina.objects.get(d_nombre=disciplina3)
            perfil.save()
            resultados['userProfile'] = ACTUALIZADO
    except:
        resultados['userProfile'] = ERROR
        return

    try:
        marca=Marca.objects.get(m_nombre=marca)
        resultados['marca'] = REFERENCIADO
    except:
        resultados['marca'] = ERROR
        return

    try:
        if Relacion.objects.filter(r_user=usuario, r_marca=marca).exists():
            relacion=Relacion.objects.get(r_user=usuario, r_marca=marca)
            relacion.r_estado='A'
            relacion.save()
            resultados['relacion'] = REFERENCIADO
        else:
            relacion=Relacion.objects.create(r_user=usuario, r_marca=marca, r_estado='A')
            resultados['relacion'] = CREADO
    except:
        resultados['relacion'] = ERROR
        return

    try:
        if Saldo.objects.filter(s_user=usuario, s_marca=marca).exists():
            saldo=Saldo.objects.get(s_user=usuario, s_marca=marca)
            saldo.s_saldo=creditos
            saldo.save()
            resultados['saldo'] = REFERENCIADO
        else:
            saldo=Saldo.objects.create(s_user=usuario, s_marca=marca, s_saldo=creditos)
            resultados['saldo'] = CREADO
    except:
        resultados['saldo'] = ERROR
        return

    try:
        if Planes.objects.filter(p_usuario=usuario,p_marca=marca,p_nombre='Inicial').exists():
            plan=Planes.objects.get(p_usuario=usuario,p_marca=marca,p_nombre='Inicial')
            plan.p_creditos_totales=creditos
            plan.p_creditos_usados=0
            plan.p_fecha_obtencion=datetime.now()
            plan.p_fecha_caducidad=datetime.now()+ timedelta(days=30)
            plan.save()
            resultados['plan'] = ACTUALIZADO
        else:
            plan=Planes.objects.create(p_usuario=usuario,p_marca=marca,p_nombre='Inicial',p_creditos_totales=creditos,p_creditos_usados=0,p_fecha_obtencion=datetime.now(),p_fecha_caducidad=datetime.now()+ timedelta(days=30))
            resultados['plan'] = CREADO
    except:
        resultados['plan'] = ERROR

    return


def trim(valor):
    if valor == None:
        return ''
    else:
        return valor.strip()


def soloNumeros(valor):
    valorNumerico = ''
    digitos = '0123456789'
    for d in valor:
        if d in digitos:
            valorNumerico += d
    return valorNumerico

def grabaResultados(usuario,workbook,resultados):
    referencia={'B':'user','C':'phone','D':'userProfile','E':'marca','F':'relacion','G':'saldo','H':'plan'}
    row=workbook['resultados'].max_row+1
    workbook['resultados']['A'+str(row)]=usuario
    for letra in 'BCDEFGH':
        workbook['resultados'][letra + str(row)] = ACCION[resultados[referencia[letra]]]
    workbook.save(filename='../../sources/dataZonaFuncional.xlsx')
    return

def grabaEncabezados(workbook):
    row=workbook['resultados'].max_row+1
    _cell = workbook['resultados'].cell('A'+str(row))
    _cell.font = Font(bold=True)
    workbook['resultados']['A' + str(row)] = 'Fecha de corrida: '+fechaCorta(datetime.now())+ ' a las ' +horaCivil(datetime.now().time())
    row=workbook['resultados'].max_row+1
    workbook['resultados']['A' + str(row)] = 'usuario'
    workbook['resultados']['B' + str(row)] = 'user'
    workbook['resultados']['C' + str(row)] = 'phone'
    workbook['resultados']['D' + str(row)] = 'userProfile'
    workbook['resultados']['E' + str(row)] = 'marca'
    workbook['resultados']['F' + str(row)] = 'relacion'
    workbook['resultados']['G' + str(row)] = 'saldo'
    workbook['resultados']['H' + str(row)] = 'plan'
    workbook.save(filename='../../sources/dataZonaFuncional.xlsx')
    return

def grabaFooter(workbook):
    row=workbook['resultados'].max_row+1
    _cell = workbook['resultados'].cell('A'+str(row))
    _cell.font = Font(bold=True)
    workbook['resultados']['A' + str(row)] = '--- fin corrida ---'
    row=workbook['resultados'].max_row+1
    workbook['resultados']['A' + str(row)] = ''
    workbook.save(filename='../../sources/dataZonaFuncional.xlsx')
    return